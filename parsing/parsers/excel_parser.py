"""
ПАРСЕР EXCEL ФАЙЛОВ
===================

Логика работы:
1. Загружает Excel файл (openpyxl)
2. Пропускает первые 3 вкладки (SKIP_FIRST_SHEETS)
3. Парсит вкладки до "УП технической разработки" (STOP_SHEET_NAME)
4. Для каждой вкладки:
   - Парсит ВЕСЬ документ до конца (включая продолжение, второй журнал со строки 65-70 и т.д.)
   - Находит ВСЕ строки заголовков (поиск "ФИО" по всему листу)
   - Для каждого найденного журнала:
     * Находит колонку с ФИО студентов
     * Находит строку с месяцами (обычно строка 5 или около заголовка)
     * Определяет месяц для каждой колонки с датами
     * Находит колонки с датами и правильно парсит их с учетом месяца
     * Извлекает данные: ФИО, дата, оценка/пропуск
     * Парсит данные до следующего заголовка или до конца листа
5. Возвращает список словарей с данными (включая данные из всех журналов)

Функции:
- parse_excel_file() - главная функция парсинга файла
- parse_sheet() - парсинг одной вкладки (поддерживает несколько журналов, парсит весь документ)
- parse_date() - парсинг даты из различных форматов
- parse_grade_value() - парсинг оценки/пропуска
- find_student_column() - поиск колонки с ФИО
- find_date_columns() - поиск колонок с датами (определяет месяц для каждой колонки)
- find_month_in_cell() - поиск названия месяца в ячейке
"""

import os
import re
from openpyxl import load_workbook
from datetime import datetime, date as date_type
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import SKIP_FIRST_SHEETS, STOP_SHEET_NAME


def normalize_fio_to_initials(fio: str) -> str:
    """
    Преобразует ФИО в формат "Фамилия И.О."
    
    Примеры:
    - "Иванов Иван Иванович" -> "Иванов И.И."
    - "Петров Петр" -> "Петров П."
    - "Сидоров С.С." -> "Сидоров С.С." (уже в нужном формате)
    - "Ельченинов Владислав Антонович" -> "Ельченинов В.А."
    """
    if not fio:
        return fio
    
    # Убираем лишние пробелы
    fio = ' '.join(fio.strip().split())
    
    # Если уже в формате "Фамилия И.О." или "Фамилия И.О", оставляем как есть
    if re.match(r'^[А-ЯЁ][а-яё]+\s+[А-ЯЁ]\.\s*[А-ЯЁ]?\.?$', fio):
        return fio
    
    # Разбиваем на части
    parts = fio.split()
    
    if len(parts) == 0:
        return fio
    
    # Фамилия - первая часть
    surname = parts[0]
    
    # Имя - вторая часть (если есть)
    if len(parts) >= 2:
        first_name_initial = parts[1][0].upper() if parts[1] else ''
    else:
        first_name_initial = ''
    
    # Отчество - третья часть (если есть)
    if len(parts) >= 3:
        last_name_initial = parts[2][0].upper() if parts[2] else ''
    else:
        last_name_initial = ''
    
    # Формируем результат
    if first_name_initial and last_name_initial:
        return f"{surname} {first_name_initial}.{last_name_initial}."
    elif first_name_initial:
        return f"{surname} {first_name_initial}."
    else:
        return surname


def parse_date(date_value):
    """
    Парсинг даты из различных форматов
    
    Поддерживает:
    - datetime объекты
    - date объекты
    - Числа Excel (номер дня с 1900-01-01)
    - Строки в форматах: %d.%m.%Y, %Y-%m-%d, %d/%m/%Y, %Y/%m/%d
    """
    if isinstance(date_value, (datetime, date_type)):
        if isinstance(date_value, datetime):
            parsed_date = date_value.date()
        else:
            parsed_date = date_value
        
        # Проверяем, что дата разумная (не 1900 год)
        if parsed_date.year < 2000:
            return None
        return parsed_date
    
    elif isinstance(date_value, (int, float)):
        # Excel хранит даты как числа (дни с 1900-01-01)
        try:
            from openpyxl.utils.datetime import from_excel
            parsed_date = from_excel(date_value).date()
            
            # Проверяем, что дата разумная (не 1900 год)
            if parsed_date.year < 2000:
                # Пробуем альтернативный способ - может быть это номер дня в другом формате
                # Если число маленькое (1-31), это может быть день месяца, а не дата Excel
                if date_value < 100:
                    return None
                return None
            return parsed_date
        except Exception as e:
            return None
    
    elif isinstance(date_value, str):
        # Пытаемся распарсить строку
        date_str = date_value.strip()
        
        # Пробуем разные форматы
        formats = [
            "%d.%m.%Y", 
            "%Y-%m-%d", 
            "%d/%m/%Y", 
            "%Y/%m/%d",
            "%d.%m.%y",  # короткий год
            "%d/%m/%y"
        ]
        
        for fmt in formats:
            try:
                parsed_date = datetime.strptime(date_str, fmt).date()
                # Проверяем, что дата разумная
                if parsed_date.year >= 2000:
                    return parsed_date
            except:
                continue
        
        # Если не получилось распарсить, возвращаем None
        return None
    
    return None


def parse_grade_value(value):
    """
    Парсинг значения оценки/пропуска
    
    Логика:
    - Если содержит слова "пропуск", "н", "н/я" и т.д. → возвращает "пропуск"
    - Иначе пытается извлечь число (оценку)
    - Исключает даты и другие некорректные значения
    - Возвращает строку с оценкой или "пропуск"
    """
    # КРИТИЧЕСКИ ВАЖНО: Строгая проверка на пустые значения
    if value is None:
        return None
    
    # Если это дата - не парсим как оценку
    if isinstance(value, (datetime, date_type)):
        return None
    
    # Преобразуем в строку и проверяем на пустоту
    value_str = str(value).strip()
    
    # Если пустая строка или только пробелы - не парсим
    if not value_str or value_str == '' or value_str.isspace():
        return None
    
    # Если это просто 0 (не оценка) - пропускаем
    if value_str == '0' or value_str == '0.0':
        return None
    
    value_lower = value_str.lower()
    
    # Исключаем даты (формат YYYY-MM-DD или похожий)
    if re.match(r'^\d{4}-\d{2}-\d{2}', value_str):
        return None
    
    # Исключаем значения, которые выглядят как даты с временем
    if re.match(r'^\d{4}-\d{2}-\d{2}\s+\d{2}', value_str):
        return None
    
    # ВАЖНО: Проверяем на дробь СНАЧАЛА, до проверки на пропуск!
    # Дроби типа "н/5", "д/4", "н/б", "3/5" и т.д.
    if '/' in value_str:
        parts = value_str.split('/')
        if len(parts) >= 2:
            first_part = parts[0].strip().lower()
            second_part = parts[1].strip().lower()
            
            # Если вторая часть - это "б" или "н", то это пропуск
            if second_part in ['б', 'н', 'нб', 'н/б']:
                return "пропуск"
            
            # Проверяем, являются ли обе части числами
            try:
                first_num = int(parts[0].strip())
                second_num = int(parts[1].strip())
                
                # Если обе части - валидные оценки (1-5), возвращаем дробь целиком
                if (1 <= first_num <= 5) and (1 <= second_num <= 5):
                    return f"{first_num}/{second_num}"
            except ValueError:
                pass
            
            # Если только вторая часть - число (для "н/5", "д/4" и т.д.)
            try:
                second_num = int(parts[1].strip())
                if 1 <= second_num <= 5:
                    # Проверяем, что первая часть - это буква (н, д и т.д.)
                    if first_part.isalpha() and len(first_part) <= 2:
                        return str(second_num)
            except ValueError:
                pass
            
            # Если только первая часть - число
            try:
                first_num = int(parts[0].strip())
                if 1 <= first_num <= 5:
                    return str(first_num)
            except ValueError:
                pass
    
    # Проверяем на пропуск (ПОСЛЕ проверки на дроби!)
    # Проверяем только если это не дробь
    if any(word == value_lower for word in ['н', 'нб', 'н/б', 'пропуск', 'н/я', 'неявка']) or value_str == '*':
        return "пропуск"
    
    # Пытаемся извлечь оценку (число)
    # Сначала пробуем простое преобразование в число
    try:
        grade_num = float(value_str)
        # Проверяем, что это разумная оценка (1-5)
        if 1 <= grade_num <= 5:
            return str(int(grade_num))
        # Или оценка в формате 0-100
        elif 0 <= grade_num <= 100:
            return str(int(grade_num))
    except ValueError:
        pass
    
    # Если не получилось, пытаемся извлечь число из строки
    numbers = re.findall(r'\d+', value_str)
    if numbers:
        grade = numbers[0]
        grade_int = int(grade)
        # Проверяем, что это разумная оценка (1-5)
        if grade in ['1', '2', '3', '4', '5']:
            return grade
        # Или оценка в формате 0-100
        elif 0 <= grade_int <= 100:
            return grade
    
    # Если это короткая строка (1-2 символа) и не дата - возвращаем как есть
    if len(value_str) <= 2:
        return value_str
    
    # Иначе игнорируем (вероятно, это не оценка)
    return None


def find_student_column(header_row):
    """
    Поиск колонки с ФИО студентов
    
    Ищет колонку, содержащую слова: "фио", "студент", "фамилия", "имя" и т.д.
    Если не находит, берет первую колонку с текстом
    """
    for idx, cell in enumerate(header_row):
        if cell.value:
            value = str(cell.value).lower().strip()
            if any(word in value for word in ['фио', 'студент', 'фамилия', 'имя', 'ученик', 'учащийся']):
                return idx
    # Если не нашли, пробуем первую колонку с текстом
    for idx, cell in enumerate(header_row):
        if cell.value and isinstance(cell.value, str) and len(cell.value.strip()) > 2:
            return idx
    return None


def find_month_in_cell(cell_value):
    """
    Ищет название месяца в значении ячейки
    
    Returns:
        (month_num, year) или None
    """
    if not cell_value:
        return None
    
    cell_str = str(cell_value).lower().strip()
    current_year = datetime.now().year
    
    # Ищем названия месяцев (полные и сокращенные)
    months = {
        'январь': 1, 'янв': 1, 'февраль': 2, 'фев': 2,
        'март': 3, 'апрель': 4, 'апр': 4,
        'май': 5, 'июнь': 6, 'июль': 7, 'август': 8, 'авг': 8,
        'сентябрь': 9, 'сен': 9, 'сент': 9,
        'октябрь': 10, 'окт': 10,
        'ноябрь': 11, 'ноя': 11,
        'декабрь': 12, 'дек': 12, 'дек абрь': 12, 'декабр': 12
    }
    
    for month_name, month_num in months.items():
        if month_name in cell_str:
            # Пытаемся найти год
            year = current_year
            # Ищем год в формате YYYY
            year_match = re.search(r'20\d{2}', cell_str)
            if year_match:
                year = int(year_match.group())
            
            return (month_num, year)
    
    return None


def find_date_columns(header_row, worksheet, header_row_idx):
    """
    Поиск колонок с датами с учетом месяцев и чисел
    
    Логика:
    1. Месяцы находятся на строке выше заголовка (header_row_idx - 1)
    2. Числа находятся на той же строке, что и заголовок (header_row_idx)
    3. Для каждой колонки определяет месяц из строки выше
    4. Использует месяц + число для построения полной даты
    5. Колонки с датами: от C (индекс 2) до AC (индекс 28)
    
    Возвращает список кортежей: (индекс_колонки, дата)
    """
    date_columns = []
    current_year = datetime.now().year
    
    # Строка с месяцами находится на 1 строку выше заголовка
    month_row_idx = header_row_idx - 1
    month_row = None
    
    if month_row_idx >= 1:
        month_row = list(worksheet[month_row_idx])
    
    # Для каждой колонки определяем месяц из строки выше
    column_months = {}  # индекс_колонки -> (month, year)
    last_month_year = None  # Последний найденный месяц (для распространения на следующие колонки)
    
    if month_row:
        # ВАЖНО: Обрабатываем ВСЕ колонки, включая пустые (для объединенных ячеек)
        # Распространяем месяц на все колонки до следующего месяца
        for idx, cell in enumerate(month_row):
            month_year = find_month_in_cell(cell.value)
            if month_year:
                column_months[idx] = month_year
                last_month_year = month_year
            elif last_month_year:
                # Если месяц не найден, но есть последний найденный месяц,
                # распространяем его на эту колонку (месяцы могут быть объединены)
                # Это важно для правильного определения дат в колонках с пустыми ячейками
                column_months[idx] = last_month_year
    
    # Колонка C имеет индекс 2 (A=0, B=1, C=2)
    # Таблица оценок идет от C и до конца (не ограничиваем жестко)
    # Но обычно это до колонки Z или дальше
    C_COLUMN_INDEX = 2
    
    # Парсим даты из заголовков (начиная с колонки C)
    # ВАЖНО: Обрабатываем ТОЛЬКО колонки с числами в заголовке
    # НЕ создаем даты для пустых ячеек!
    last_date = None  # Последняя найденная дата
    last_date_col_idx = None  # Индекс последней колонки с датой
    dates_in_month = {}  # месяц -> список дат (для проверки логики)
    
    for idx, cell in enumerate(header_row):
        # Начинаем парсить с колонки C
        # Не ограничиваем конец, чтобы захватить все даты
        if idx < C_COLUMN_INDEX:
            continue
        
        # КРИТИЧЕСКИ ВАЖНО: Пропускаем пустые ячейки!
        # Если ячейка пустая, значит в этой колонке нет даты
        if not cell.value:
            continue
        
        cell_value = cell.value
        current_date = None
        
        # Если это число от 1 до 31 - это день месяца
        # КРИТИЧЕСКИ ВАЖНО: Проверяем, что это именно целое число дня, а не дробное или 0
        if isinstance(cell_value, (int, float)):
            # Проверяем, что это целое число от 1 до 31 (НЕ 0, НЕ дробное)
            day_float = float(cell_value)
            day_int = int(day_float)
            
            # Если это не целое число (например, 12.5) или 0 - пропускаем
            if day_float != day_int or day_int < 1 or day_int > 31:
                continue
            
            day = day_int
            
            # Определяем месяц для этой колонки
            month_year = column_months.get(idx)
            
            # Создаем дату ТОЛЬКО если есть месяц для этой колонки
            if month_year:
                month, year = month_year
                try:
                    current_date = date_type(year, month, day)
                    
                    # ДОПОЛНИТЕЛЬНАЯ ПРОВЕРКА: Проверяем, что дата логична
                    # Если есть предыдущая дата в том же месяце, проверяем разницу
                    if month not in dates_in_month:
                        dates_in_month[month] = []
                    
                    # Если в месяце уже есть даты, проверяем логику
                    if dates_in_month[month]:
                        # Если новая дата меньше последней даты в месяце более чем на 7 дней,
                        # это может быть ошибка (дата из другого места таблицы)
                        last_date_in_month = max(dates_in_month[month])
                        days_diff = (current_date - last_date_in_month).days
                        
                        # Если разница отрицательная (дата раньше последней) и больше 7 дней,
                        # это подозрительно - возможно, это дата из другой части таблицы
                        if days_diff < -7:
                            # Пропускаем эту дату - она выглядит как ошибка
                            continue
                    
                    dates_in_month[month].append(current_date)
                    last_date = current_date
                    last_date_col_idx = idx
                except ValueError:
                    # Если дата невалидна (например, 31 февраля), пропускаем
                    pass
        
        # Пытаемся распарсить как полную дату (на случай если формат другой)
        if not current_date:
            parsed_date = parse_date(cell_value)
            if parsed_date:
                current_date = parsed_date
                last_date = current_date
                last_date_col_idx = idx
        
        # Добавляем дату в список, если она найдена
        if current_date:
            date_columns.append((idx, current_date))
    
    return date_columns


def parse_topics_table(worksheet, group_name, subject_name, start_row=1):
    """
    Парсит таблицу с темами занятий (Форма 1, где в колонке "Кол-во часов" всегда "2")
    
    Ищет таблицу с заголовками:
    - "Дата проведения"
    - "Кол-во часов" 
    - "Наименование учебного занятия"
    
    Фильтрация:
    - Не парсит строки после колонки AE (индекс 30), где:
      * "Кол-во часов" = 2 (или "2")
      * В названии есть "Тема"
      * В названии есть "4" (как часть темы, например "Тема 1.1. ... 4. ...")
    
    Возвращает список тем занятий
    """
    topics_data = []
    max_row = worksheet.max_row
    max_col = worksheet.max_column
    
    # Колонка AE имеет индекс 30 (A=0, B=1, ..., AE=30)
    AE_COLUMN_INDEX = 30
    
    # Ищем заголовок таблицы с темами
    header_row_idx = None
    date_col = None
    hours_col = None
    topic_col = None
    
    # Ищем строку с заголовками "Кол-во часов" или "Наименование учебного занятия"
    # Ищем по всему листу, не только в первых 100 строках
    for row_idx in range(start_row, max_row + 1):
        row = list(worksheet[row_idx])
        row_text_lower = ' '.join([str(cell.value).lower() if cell.value else '' for cell in row])
        
        # Ищем ключевые слова заголовков - более гибкий поиск
        has_hours = 'кол-во часов' in row_text_lower or 'количество часов' in row_text_lower or 'часов' in row_text_lower
        has_topic = 'наименование учебного занятия' in row_text_lower or ('наименование' in row_text_lower and 'занятия' in row_text_lower) or 'наименование' in row_text_lower
        
        if has_hours and has_topic:
            header_row_idx = row_idx
            
            # Находим колонки
            for col_idx, cell in enumerate(row):
                if not cell.value:
                    continue
                cell_value = str(cell.value).lower().strip()
                
                if 'дата проведения' in cell_value:
                    date_col = col_idx
                elif 'кол-во часов' in cell_value or 'количество часов' in cell_value or ('часов' in cell_value and 'кол' in cell_value):
                    hours_col = col_idx
                elif 'наименование учебного занятия' in cell_value or ('наименование' in cell_value and 'занятия' in cell_value):
                    topic_col = col_idx
            
            if hours_col is not None and topic_col is not None:
                break
    
    if not header_row_idx or hours_col is None or topic_col is None:
        return topics_data
    
    # Проверяем, что таблица тем находится в колонках AE и дальше
    # (колонка AE = индекс 30)
    # Таблица тем должна начинаться с колонки AE или позже
    # Проверяем, что хотя бы одна из колонок (hours_col или topic_col) находится в AE+
    if hours_col < AE_COLUMN_INDEX + 1 and topic_col < AE_COLUMN_INDEX + 1:
        # Таблица тем не в колонках AE+, пропускаем
        return topics_data
    
    # Парсим строки с темами (начинаем со следующей строки после заголовка)
    for row_idx in range(header_row_idx + 1, max_row + 1):
        row = list(worksheet[row_idx])
        
        if len(row) <= max(hours_col, topic_col):
            continue
        
        hours_cell = row[hours_col] if hours_col < len(row) else None
        topic_cell = row[topic_col] if topic_col < len(row) else None
        
        if not topic_cell or not topic_cell.value:
            continue
        
        topic_name = str(topic_cell.value).strip()
        
        # Пропускаем пустые строки и заголовки
        if not topic_name or len(topic_name) < 3:
            continue
        
        # Проверяем, что это не заголовок
        if any(keyword in topic_name.lower() for keyword in ['наименование', 'занятия', 'дата', 'кол-во', 'часов']):
            continue
        
        # Получаем количество часов (обычно "2")
        hours_value = None
        if hours_cell and hours_cell.value:
            try:
                hours_value = int(float(str(hours_cell.value).strip()))
            except:
                hours_value = str(hours_cell.value).strip()
        
        # ФИЛЬТРАЦИЯ: Не парсим строки после колонки AE, где:
        # - "Кол-во часов" = 2 (или "2")
        # - В названии есть "Тема"
        # - В названии есть "4" (как часть темы, например "Тема 1.1. ... 4. ...")
        topic_lower = topic_name.lower()
        has_tema = 'тема' in topic_lower
        has_four = '4' in topic_name or '.4.' in topic_lower or ' 4.' in topic_lower
        
        # Проверяем, что hours = 2 (может быть int 2 или строка "2")
        hours_is_two = False
        if hours_value is not None:
            if isinstance(hours_value, int) and hours_value == 2:
                hours_is_two = True
            elif isinstance(hours_value, str) and str(hours_value).strip() == "2":
                hours_is_two = True
        
        # Если все условия выполнены - пропускаем эту строку
        if hours_is_two and has_tema and has_four:
            continue
        
        # Получаем дату проведения (если есть)
        date_value = None
        if date_col is not None and date_col < len(row):
            date_cell = row[date_col]
            if date_cell and date_cell.value:
                parsed_date = parse_date(date_cell.value)
                if parsed_date:
                    date_value = parsed_date
        
        # Добавляем тему (даже если часов нет, но есть название темы)
        topics_data.append({
            'group': group_name,
            'subject': subject_name,
            'topic': topic_name,
            'hours': hours_value if hours_value else 2,  # По умолчанию 2, как указал пользователь
            'date': date_value
        })
    
    return topics_data


def parse_sheet(worksheet, group_name, subject_name):
    """
    Парсинг одного листа Excel с поддержкой нескольких журналов
    
    Логика:
    1. Парсит ВЕСЬ документ до конца (включая продолжение, второй журнал со строки 65-70 и т.д.)
    2. Ищет все строки заголовков (ищет "ФИО" по всему листу)
    3. Для каждого найденного заголовка:
       - Находит колонку с ФИО
       - Находит колонки с датами (определяет месяц для каждой колонки из строки с месяцами)
       - Парсит данные до следующего заголовка или до конца листа
    4. Также парсит таблицу с темами занятий (где "Кол-во часов" = 2)
    5. Возвращает список словарей с данными (все данные из всех журналов + темы)
    """
    data = []
    max_row = worksheet.max_row
    
    # Находим все строки с заголовками "ФИО" по всему листу
    header_rows = []
    for idx in range(1, max_row + 1):
        row = list(worksheet[idx])
        for cell in row:
            if cell.value:
                cell_value = str(cell.value).lower().strip()
                if 'фио' in cell_value or 'студент' in cell_value:
                    header_rows.append(idx)
                    break
    
    if not header_rows:
        return data
    
    # Список заголовков, которые не являются студентами
    header_keywords = [
        'месяц/число', 'фио обучающихся', 'фио', 'кол-во часов', 
        'количество часов', 'часы', 'студент', 'обучающийся',
        'фамилия', 'имя', 'отчество', 'дата', 'оценка', 'пропуск'
    ]
    
    def is_header_row(fio_text):
        """Проверяет, является ли строка заголовком, а не студентом"""
        fio_lower = fio_text.lower().strip()
        # Если содержит ключевые слова заголовков
        if any(keyword in fio_lower for keyword in header_keywords):
            return True
        # Если слишком короткое или содержит только цифры/символы
        if len(fio_lower) < 3:
            return True
        # Если содержит только цифры или специальные символы
        if fio_lower.replace(' ', '').replace('.', '').replace('-', '').isdigit():
            return True
        return False
    
    # Парсим каждый найденный журнал
    for journal_idx, header_row_idx in enumerate(header_rows):
        # Определяем границы журнала
        # Начало: строка после заголовка
        start_row = header_row_idx + 1
        # Конец: следующий заголовок или конец листа
        if journal_idx + 1 < len(header_rows):
            end_row = header_rows[journal_idx + 1]
        else:
            end_row = max_row + 1
        
        # Получаем строку заголовков
        header_row = list(worksheet[header_row_idx])
        student_col = find_student_column(header_row)
        date_columns = find_date_columns(header_row, worksheet, header_row_idx)
        
        if student_col is None:
            continue
        
        if not date_columns:
            continue
        
        # Парсим данные студентов в этом журнале
        for row_idx in range(start_row, end_row):
            row = list(worksheet[row_idx])
            
            if len(row) <= student_col:
                continue
            
            student_cell = row[student_col]
            if not student_cell or not student_cell.value:
                continue
            
            student_fio = str(student_cell.value).strip()
            
            # Пропускаем заголовки (на случай, если они повторяются)
            if is_header_row(student_fio):
                continue
            
            if not student_fio or len(student_fio) < 3:  # Минимум 3 символа для ФИО
                continue
            
            # Нормализуем ФИО (убираем лишние пробелы)
            student_fio = ' '.join(student_fio.split())
            
            # Преобразуем в формат "Фамилия И.О." если это полное ФИО
            student_fio = normalize_fio_to_initials(student_fio)
            
            # Парсим оценки для каждой даты
            # ВАЖНО: Берем данные ТОЛЬКО из колонок, которые были найдены в find_date_columns
            # Не создаем даты для колонок без чисел в заголовке
            for date_col_idx, date in date_columns:
                # Проверяем, что строка достаточно длинная
                if len(row) <= date_col_idx:
                    continue
                
                cell = row[date_col_idx]
                
                # КРИТИЧЕСКИ ВАЖНО: Пропускаем пустые ячейки
                # Если ячейка пустая (None, пустая строка, пробелы), не парсим оценку
                if not cell or cell.value is None:
                    continue
                
                # Проверяем, что значение не пустая строка
                if isinstance(cell.value, str) and not cell.value.strip():
                    continue
                
                # Парсим оценку только если ячейка не пустая
                grade_value = parse_grade_value(cell.value)
                
                # Добавляем оценку только если она валидна и не None
                if grade_value:
                    data.append({
                        'group': group_name,
                        'subject': subject_name,
                        'fio': student_fio,
                        'date': date,
                        'grade': grade_value
                    })
    
    # Парсим таблицу с темами занятий (ищем по всему листу)
    topics_data = parse_topics_table(worksheet, group_name, subject_name, start_row=1)
    
    # Добавляем темы к данным (сохраняем как специальный тип данных)
    for topic in topics_data:
        data.append({
            'group': group_name,
            'subject': subject_name,
            'type': 'topic',  # Помечаем как тему
            'topic': topic['topic'],
            'hours': topic['hours'],
            'date': topic.get('date')
        })
    
    return data


def calculate_subject_statistics(sheet_data):
    """
    Вычисляет статистику по предмету из распарсенных данных
    
    Возвращает словарь с:
    - total_classes: общее количество занятий (уникальных дат)
    - total: общее количество записей (оценок + пропусков)
    - grades_count: количество оценок (не пропусков)
    - absences_count: количество пропусков
    - attendance_percent: процент посещаемости (0-100)
    """
    total = 0
    grades_count = 0
    absences_count = 0
    unique_dates = set()
    
    # Фильтруем только записи с оценками/пропусками (не темы и не статистику)
    grade_records = [item for item in sheet_data 
                     if item.get('type') not in ['topic', 'statistics'] 
                     and item.get('grade')]
    
    for item in grade_records:
        grade_value = item.get('grade', '')
        date = item.get('date')
        
        # Добавляем дату в набор уникальных дат
        if date:
            unique_dates.add(date)
        
        if grade_value:
            total += 1
            if grade_value.lower() in ['пропуск', 'н', 'н/я', 'неявка', 'нб', 'н/б'] or grade_value == '*':
                absences_count += 1
            else:
                grades_count += 1
    
    # Общее количество занятий = количество уникальных дат
    total_classes = len(unique_dates)
    
    # Вычисляем процент посещаемости
    # Посещаемость = (количество оценок / общее количество) * 100
    # Если нет записей, посещаемость = 0%
    if total > 0:
        attendance_percent = round((grades_count / total) * 100, 1)
    else:
        attendance_percent = 0.0
    
    return {
        'total_classes': total_classes,
        'total': total,
        'grades_count': grades_count,
        'absences_count': absences_count,
        'attendance_percent': attendance_percent
    }


def parse_excel_file(file_path):
    """
    Парсинг Excel файла
    
    Логика:
    1. Загружает файл
    2. Извлекает название группы из имени файла
    3. Находит первую вкладку, начинающуюся с "ОГСЭ" или "ОГЭ"
    4. Парсит все вкладки до тех, которые начинаются с "УП"
    5. Для каждой вкладки вызывает parse_sheet() (даже если там нет данных)
    6. Вычисляет статистику для каждого предмета
    7. Возвращает объединенные данные с информацией о статистике
    """
    
    try:
        # Загружаем файл с data_only=True для получения вычисленных значений
        # Но даты будем парсить специальным образом
        workbook = load_workbook(file_path, data_only=True)
        sheet_names = workbook.sheetnames
        
        # Извлекаем название группы из имени файла
        filename = os.path.basename(file_path)  # Получаем только имя файла
        group_name = filename.replace('Испп ', '').replace('.xslm', '').replace('.xlsx', '').replace('temp_', '')
        
        all_data = []
        
        # Находим первую вкладку, начинающуюся с "ОГСЭ" или "ОГЭ"
        start_idx = None
        for idx, sheet_name in enumerate(sheet_names):
            sheet_name_upper = sheet_name.upper().strip()
            if sheet_name_upper.startswith('ОГСЭ') or sheet_name_upper.startswith('ОГЭ'):
                start_idx = idx
                break
        
        # Если не нашли вкладку с ОГСЭ/ОГЭ, используем старую логику (пропускаем первые 3)
        if start_idx is None:
            start_idx = min(SKIP_FIRST_SHEETS, len(sheet_names))
        
        # Находим индекс первой вкладки, начинающейся с "УП"
        end_idx = len(sheet_names)
        for idx, sheet_name in enumerate(sheet_names):
            sheet_name_upper = sheet_name.upper().strip()
            if sheet_name_upper.startswith('УП'):
                end_idx = idx
                break
        
        # Если не нашли вкладку с УП, используем старую логику
        if end_idx == len(sheet_names):
            for idx, sheet_name in enumerate(sheet_names):
                if STOP_SHEET_NAME.lower() in sheet_name.lower():
                    end_idx = idx
                    break
        
        if start_idx is None or start_idx >= end_idx:
            workbook.close()
            return all_data
        
        # Парсим все вкладки от ОГСЭ до УП (даже если там нет данных)
        for idx in range(start_idx, end_idx):
            sheet_name = sheet_names[idx]
            worksheet = workbook[sheet_name]
            
            # Извлекаем название предмета из названия вкладки
            subject_name = sheet_name
            
            # Парсим вкладку (даже если там нет данных, parse_sheet вернет пустой список)
            sheet_data = parse_sheet(worksheet, group_name, subject_name)
            
            # Вычисляем статистику для этого предмета
            statistics = calculate_subject_statistics(sheet_data)
            
            # Добавляем статистику к данным предмета
            # Сохраняем статистику как специальную запись типа 'statistics'
            all_data.append({
                'group': group_name,
                'subject': subject_name,
                'type': 'statistics',
                'total_classes': statistics['total_classes'],
                'total': statistics['total'],
                'grades_count': statistics['grades_count'],
                'absences_count': statistics['absences_count'],
                'attendance_percent': statistics['attendance_percent']
            })
            
            # Добавляем все остальные данные (оценки, темы и т.д.)
            all_data.extend(sheet_data)
        
        workbook.close()
        return all_data
        
    except Exception as e:
        print(f"Ошибка при парсинге файла {file_path}: {e}")
        import traceback
        traceback.print_exc()
        return []

